"""Bulk employee import — parse → validate (dry-run) → commit.

See EMPLOYEE-IMPORT-PLAN.md. Direct-create model: each valid row creates a
User + PrivateUser + Job + Salary + a new-model BASIC structure assignment, so
imported employees are payroll-ready and identical to seeder/wizard-created ones.

Pure-ish: parse() and validate() never write. commit() writes and is
idempotent (existing email/passport → skipped, never duplicated). No HTTP here.
"""
from __future__ import annotations

import csv
import io
import re
import secrets
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from core.security import generate_passwd_hash
from core.model import (
    Department, EmployeeSalaryAssignment, Job, PrivateUser, Salary,
    SalaryComponent, SalaryStructure, SalaryStructureLine, User, UserType,
)

# ── Schema ──────────────────────────────────────────────────────────────────
REQUIRED_COLUMNS = [
    "first_name", "last_name", "email", "job_title", "start_date",
    "base_salary", "currency",
]
RECOMMENDED_COLUMNS = [
    "passport_number", "department", "work_days_per_week", "hours_per_month",
    "role", "pay_basis", "work_start_time", "work_end_time",
]
OPTIONAL_COLUMNS = [
    "dob", "nationality", "permit_type", "permit_number", "permit_expiry",
    "has_permission_to_work", "deduct_transport", "deduct_accommodation", "notes",
]
TEMPLATE_COLUMNS = REQUIRED_COLUMNS + RECOMMENDED_COLUMNS + OPTIONAL_COLUMNS

ALLOWED_CURRENCIES = {"MUR", "MGA", "TZS", "USD", "EUR", "GBP"}
# Monthly statutory minima used as a WARNING (not a hard block) — mirrors the
# wizard's MIN_WAGES. Only currencies with a known floor are checked.
MIN_WAGE_MONTHLY = {"MUR": Decimal("11275"), "MGA": Decimal("258960")}
ALLOWED_ROLES = {"employee", "manager", "admin", "owner"}
ALLOWED_PAY_BASIS = {"monthly", "hourly", "daily"}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _company_currency(db: Session, company_id: int) -> Optional[str]:
    """The company's operating currency, derived from its country
    (Company.currency → Country.currency). Currency is server-authoritative:
    every employee is paid in their company's country currency, so the CSV /
    payload `currency` column is advisory only and is overridden by this on
    write. Returns None only if the company/country can't be resolved, in which
    case callers fall back to the client value (last resort) or the model
    default. Mirrors crud/job.resolve_salary_currency for the direct-write path."""
    from core.model import Company
    company = db.query(Company).filter(Company.company_id == company_id).first()
    return getattr(company, "currency", None) if company else None


def _norm_header(h: Any) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def _s(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


# ── Parse ───────────────────────────────────────────────────────────────────
def parse(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    """Parse a CSV or XLSX upload into normalized row dicts (header→value).
    Raises ValueError on an unreadable/empty file."""
    name = (filename or "").lower()
    rows: List[Dict[str, str]] = []
    if name.endswith(".csv") or (not name.endswith(".xlsx") and not name.endswith(".xls")):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV has no header row.")
        fieldmap = {fn: _norm_header(fn) for fn in reader.fieldnames}
        for raw in reader:
            rows.append({fieldmap[k]: _s(v) for k, v in raw.items() if k in fieldmap})
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise ValueError("openpyxl is required to read .xlsx files") from e
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = [_norm_header(h) for h in next(it)]
        except StopIteration:
            raise ValueError("Spreadsheet is empty.")
        for raw in it:
            if raw is None or all(c is None or _s(c) == "" for c in raw):
                continue  # skip blank rows
            rows.append({header[i]: _s(raw[i]) for i in range(min(len(header), len(raw)))})
    return rows


# ── Validate (no writes) ────────────────────────────────────────────────────
def validate(db: Session, company_id: int, rows: List[Dict[str, str]]) -> Dict[str, Any]:
    """Validate parsed rows against the schema + existing data. Returns
    {total, ok:[{row,data}], errors:[{row,field,reason}], warnings:[{row,reason}],
     duplicates_in_file, missing_columns}. Writes nothing."""
    present_cols = set().union(*[set(r.keys()) for r in rows]) if rows else set()
    missing_columns = [c for c in REQUIRED_COLUMNS if c not in present_cols]

    # Existing keys for dup detection (one query each).
    existing_emails = {e.lower() for (e,) in db.query(User.email).all() if e}
    existing_passports = {
        p.lower() for (p,) in db.query(PrivateUser.pass_port_number).all() if p
    }

    # Server-authoritative operating currency (from the company's country). The
    # per-row `currency` column is advisory — validation and write both use this.
    company_currency = _company_currency(db, company_id)

    ok: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    seen_emails: Dict[str, int] = {}
    seen_passports: Dict[str, int] = {}

    for i, r in enumerate(rows):
        rownum = i + 2  # +1 for 0-index, +1 for header row → spreadsheet row number
        row_errors: List[Tuple[str, str]] = []

        email = _s(r.get("email")).lower()
        if not email:
            row_errors.append(("email", "missing"))
        elif not EMAIL_RE.match(email):
            row_errors.append(("email", "invalid format"))
        for f in ("first_name", "last_name", "job_title"):
            if not _s(r.get(f)):
                row_errors.append((f, "missing"))

        # start_date
        sd = _s(r.get("start_date"))
        if not sd:
            row_errors.append(("start_date", "missing"))
        elif _parse_date(sd) is None:
            row_errors.append(("start_date", "not a valid YYYY-MM-DD date"))

        # currency — server-authoritative: derived from the company's country,
        # NOT the CSV column. The client value is ignored (kept only as a last
        # resort if the company currency can't be resolved).
        currency = (company_currency or _s(r.get("currency")).upper() or "MUR").upper()
        if currency not in ALLOWED_CURRENCIES:
            row_errors.append(("currency", f"unsupported company currency '{currency}'"))

        # base_salary
        salary = _parse_decimal(_s(r.get("base_salary")))
        if _s(r.get("base_salary")) == "":
            row_errors.append(("base_salary", "missing"))
        elif salary is None or salary <= 0:
            row_errors.append(("base_salary", "not a positive number"))

        # role / pay_basis (optional but validated when present)
        role = _s(r.get("role")).lower() or "employee"
        if role not in ALLOWED_ROLES:
            row_errors.append(("role", f"must be one of {', '.join(sorted(ALLOWED_ROLES))}"))
        pay_basis = _s(r.get("pay_basis")).lower() or "monthly"
        if pay_basis not in ALLOWED_PAY_BASIS:
            row_errors.append(("pay_basis", f"must be one of {', '.join(sorted(ALLOWED_PAY_BASIS))}"))

        # duplicates
        if email:
            if email in existing_emails:
                row_errors.append(("email", "already exists — will be skipped"))
            elif email in seen_emails:
                row_errors.append(("email", f"duplicate of row {seen_emails[email]} in this file"))
            else:
                seen_emails[email] = rownum
        passport = _s(r.get("passport_number")).lower()
        if passport:
            if passport in existing_passports:
                row_errors.append(("passport_number", "already exists — will be skipped"))
            elif passport in seen_passports:
                row_errors.append(("passport_number", f"duplicate of row {seen_passports[passport]} in this file"))
            else:
                seen_passports[passport] = rownum

        if row_errors:
            for field, reason in row_errors:
                errors.append({"row": rownum, "field": field, "reason": reason})
            continue

        # WARNINGS (don't block) — below minimum wage.
        floor = MIN_WAGE_MONTHLY.get(currency)
        if pay_basis == "monthly" and floor is not None and salary is not None and salary < floor:
            warnings.append({
                "row": rownum,
                "reason": f"{currency} {salary} is below the statutory monthly minimum ({currency} {floor})",
            })

        # WARNINGS — a work time was given but couldn't be parsed: it'll be left
        # blank (no clock reminders) rather than silently misread. Use HH:MM.
        for tf in ("work_start_time", "work_end_time"):
            raw = _s(r.get(tf))
            if raw and _parse_time(raw) is None:
                warnings.append({
                    "row": rownum,
                    "reason": f"{tf} '{raw}' is not a valid time (use HH:MM, e.g. 08:00) — will be left blank",
                })

        ok.append({"row": rownum, "data": r})

    return {
        "total": len(rows),
        "ok": ok,
        "errors": errors,
        "warnings": warnings,
        "missing_columns": missing_columns,
    }


# ── Commit (writes) ─────────────────────────────────────────────────────────
def commit(db: Session, company_id: int, rows: List[Dict[str, str]], actor_user_id: Optional[int]) -> Dict[str, Any]:
    """Create employees for every VALID row. Re-validates first (never trust the
    client). Idempotent: existing email/passport rows are skipped, not
    duplicated. Returns {created, skipped, failed:[...]}. Caller commits."""
    result = validate(db, company_id, rows)
    if result["missing_columns"]:
        raise ValueError(f"Missing required columns: {', '.join(result['missing_columns'])}")

    from api.v1.account_claim import generate_claim_token
    from core.model import Company

    company = db.query(Company).filter(Company.company_id == company_id).first()
    # showUser (the login response) requires jobs[].employer_name/employer_brn as
    # non-null strings — without these an imported employee 500s on login.
    employer_name = (company.company_name if company else "") or ""
    employer_brn = (company.brn if company else "") or ""

    dept_cache: Dict[str, Department] = {}
    basic_component = _get_or_create_basic_component(db, company_id)
    # Server-authoritative currency for every row (company's country currency).
    company_currency = _company_currency(db, company_id)

    created = 0
    claims: List[Dict[str, str]] = []  # {email, name, token} so the employer can hand out claim links
    failed: List[Dict[str, Any]] = []

    for item in result["ok"]:
        r = item["data"]
        rownum = item["row"]
        try:
            email = _s(r.get("email")).lower()
            # Idempotency guard (race-safe within this txn).
            if db.query(User).filter(User.email == email).first() is not None:
                continue

            user = User(
                user_type=UserType.private if hasattr(UserType, "private") else "private",
                email=email, user_name=email,
                password_hash=generate_passwd_hash(secrets.token_urlsafe(24)),
                onboard_complete=True, user_verified=False, user_enabled=True,
                company_onboarding_status="approved",
            )
            db.add(user)
            db.flush()

            dept = _get_or_create_department(db, company_id, _s(r.get("department")), dept_cache)
            emp = PrivateUser(
                user_id=user.user_id,
                first_name=_s(r.get("first_name")), last_name=_s(r.get("last_name")),
                company_id=company_id,
                # Set on PrivateUser (not just Job) — the departments page counts
                # members by PrivateUser.department_id; Job.department_id alone
                # showed imported staff as unassigned.
                department_id=dept.department_id if dept else None,
                pass_port_number=(_s(r.get("passport_number")) or None),
                role=_s(r.get("role")).lower() or "employee",
            )
            db.add(emp)
            db.flush()
            from services.employee_code_service import ensure_employee_code
            ensure_employee_code(db, emp)

            start = _parse_date(_s(r.get("start_date")))
            wdays = _parse_int(r.get("work_days_per_week")) or 5
            wstart = _parse_time(r.get("work_start_time"))
            wend = _parse_time(r.get("work_end_time"))
            # Per-day schedule value: a 'HH:MM-HH:MM' range when both times are
            # given (matches the seeder/profile shape), else bare normal hours.
            daily_val = (
                f"{wstart.strftime('%H:%M')}-{wend.strftime('%H:%M')}"
                if wstart and wend else "8"
            )
            job = Job(
                private_user_id=emp.private_user_id, company_id=company_id,
                job_title=_s(r.get("job_title")),
                employer_name=employer_name, employer_brn=employer_brn,
                department_id=dept.department_id if dept else None,
                first_date_of_employment=datetime.combine(start, datetime.min.time()) if start else None,
                # Materialise a real schedule from work_days_per_week instead of
                # leaving {} (which makes the engine silently assume Mon–Fri).
                work_days=_build_work_days(wdays, daily_val),
                # Clock in/out reminders + tight auto-close need these (NULL → off).
                work_start_time=wstart, work_end_time=wend,
                weekly_rest_day_dow=7,
                verification_status="approved",
            )
            db.add(job)
            db.flush()

            salary_amt = _parse_decimal(_s(r.get("base_salary"))) or Decimal("0")
            # Currency is server-authoritative (company country), not the client
            # column — consistent with crud/job.create_salary / onboard_job.
            currency = (company_currency or _s(r.get("currency")).upper() or "MUR").upper()
            pay_basis = _s(r.get("pay_basis")).lower() or "monthly"
            hpm = _s(r.get("hours_per_month")) or "195"
            salary_obj = Salary(
                job_id=job.job_id, pay_basis=pay_basis, salary=str(salary_amt),
                currency=currency, monthly_hours=hpm,
                break_in_minutes_per_day=0,
                days_of_work_per_month=_days_of_work_per_month(wdays),
                # Denormalized RLS tenant column — set from the company so we
                # don't depend on the DB trigger (missing on prod's create_all
                # schema → NULL → not-null violation on insert).
                company_id=company_id,
            )
            # Keep the salary money invariant (revenue = salary + allowance) — the
            # single helper every app write path uses — so imported rows are not
            # left with a NULL `revenue` mirror, which read paths that rely on it
            # would render as 0.
            from db_models.crud.job import _enforce_salary_money
            _enforce_salary_money(salary_obj)
            db.add(salary_obj)

            # New-model BASIC structure assignment so monthly payroll resolves.
            # Per-employee structure is deliberate: the salary-assignment UI
            # assigns a structure with NO amount input (overrides:[]), so the
            # amount must live ON the structure — and editing the amount happens
            # by editing this structure. A shared template would resolve to Rs 0.
            structure = SalaryStructure(
                company_id=company_id, name=f"Imported — {emp.first_name} {emp.last_name}".strip()[:255],
                description="Created by bulk import",
            )
            db.add(structure)
            db.flush()
            db.add(SalaryStructureLine(
                structure_id=structure.id, component_id=basic_component.id,
                amount=salary_amt, order_index=0,
            ))
            db.add(EmployeeSalaryAssignment(
                private_user_id=emp.private_user_id, structure_id=structure.id,
                currency=currency, effective_from=start or datetime.now(timezone.utc).date(),
                notes="bulk import",
            ))
            db.flush()
            # One-time claim token so the worker can set a password + verify
            # (they're created unverified; login is blocked until claimed).
            token = generate_claim_token(db, user)
            claims.append({
                "email": user.email,
                "name": f"{emp.first_name} {emp.last_name}".strip(),
                "token": token,
            })
            created += 1
        except Exception as e:  # noqa: BLE001 — one bad row must not sink the batch
            failed.append({"row": rownum, "error": str(e)})

    return {
        "created": created,
        "skipped": len({e["row"] for e in result["errors"]}),
        "failed": failed,
        "warnings": result["warnings"],
        "claims": claims,
    }


# ── helpers ─────────────────────────────────────────────────────────────────
def _get_or_create_basic_component(db: Session, company_id: int) -> SalaryComponent:
    c = (
        db.query(SalaryComponent)
        .filter(SalaryComponent.company_id == company_id, SalaryComponent.is_basic == True)  # noqa: E712
        .first()
    )
    if c is None:
        c = SalaryComponent(
            company_id=company_id, code="BASIC", label="Basic salary",
            kind="earning", category="earning.basic", is_basic=True, is_taxable=True,
        )
        db.add(c)
        db.flush()
    return c


def _get_or_create_department(db: Session, company_id: int, name: str, cache: Dict[str, Department]) -> Optional[Department]:
    name = (name or "").strip()
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    d = (
        db.query(Department)
        .filter(Department.company_id == company_id, Department.name.ilike(name))
        .first()
    )
    if d is None:
        d = Department(company_id=company_id, name=name)
        db.add(d)
        db.flush()
    cache[key] = d
    return d


def _parse_date(s: str):
    s = _s(s)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _parse_time(s: str):
    """Parse a clock time from the CSV ('08:00', '8:00', '08:00:00', '8'). Returns
    a datetime.time or None. Used for Job.work_start_time / work_end_time."""
    s = _s(s)
    if s == "":
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%H"):
        try:
            return datetime.strptime(s, fmt).time()
        except (ValueError, TypeError):
            continue
    return None


def _parse_decimal(s: str) -> Optional[Decimal]:
    s = _s(s).replace(",", "")
    if s == "":
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_int(v: Any) -> Optional[int]:
    s = _s(v)
    try:
        return int(float(s)) if s else None
    except (ValueError, TypeError):
        return None


# ── Work schedule ─────────────────────────────────────────────────────────────
_WEEK_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _build_work_days(days_per_week: int, daily_hours: str = "8") -> Dict[str, str]:
    """Turn a `work_days_per_week` count into the JSONB schedule the payroll
    engine reads. Without this the engine silently falls back to a Mon–Fri week
    (`_DEFAULT_WORKDAYS`), so a 6-day worker would be prorated/absence-docked as
    if they worked 5 — wrong numbers, no error. We materialise an explicit
    schedule (first N of Mon→Sun) so the assumption is conscious, not implicit.
    Values are per-day normal hours as a string ('off'/'0' would mark a day OFF,
    so we never store those — see time_log_service._is_workday."""
    n = max(1, min(7, days_per_week))
    return {day: daily_hours for day in _WEEK_ORDER[:n]}


def _days_of_work_per_month(days_per_week: int) -> int:
    """Average scheduled working days in a month for an N-day week
    (N × 52 / 12). 5→22, 6→26, 7→30. Feeds the daily-rate denominator;
    hard-coding 22 over-paid 6-day workers on the daily/absence path."""
    n = max(1, min(7, days_per_week))
    return round(n * 52 / 12)


def template_csv() -> str:
    """The downloadable CSV template: header + one example row."""
    example = {
        "first_name": "Jane", "last_name": "Doe", "email": "jane.doe@example.com",
        "job_title": "Sales Associate", "start_date": "2026-01-15",
        "base_salary": "25000", "currency": "MUR", "passport_number": "A1234567",
        "department": "Sales", "work_days_per_week": "5", "hours_per_month": "195",
        "role": "employee", "pay_basis": "monthly",
        "work_start_time": "08:00", "work_end_time": "16:00",
    }
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=TEMPLATE_COLUMNS)
    w.writeheader()
    w.writerow({c: example.get(c, "") for c in TEMPLATE_COLUMNS})
    return buf.getvalue()
