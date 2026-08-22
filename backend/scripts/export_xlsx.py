#!/usr/bin/env python3
"""Export mauritius_sector_rates.csv to a formatted Excel template."""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
csv_path = os.path.join(base, "mauritius_sector_rates.csv")
out_path = os.path.join(base, "mauritius_sector_rates_template.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Sector Rates"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Header row ───────────────────────────────────────────────────────────────
headers = [
    "Sector", "Category of Employee", "Grade / Service Details",
    "Year of Service", "Rate Type", "Rate", "Effective From", "Notes",
]
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, border

col_widths = [30, 34, 32, 22, 14, 12, 18, 36]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 32

# ── Data rows ─────────────────────────────────────────────────────────────────
fill_even = PatternFill("solid", fgColor="EBF3FB")
fill_odd  = PatternFill("solid", fgColor="FFFFFF")
data_font = Font(name="Calibri", size=10)
data_align = Alignment(vertical="center")

with open(csv_path, newline="", encoding="utf-8") as f:
    for r_idx, row in enumerate(csv.DictReader(f), 2):
        fill = fill_even if r_idx % 2 == 0 else fill_odd
        rate_val = float(row["Rate"]) if row["Rate"].strip() else ""
        values = [
            row["Sector"], row["Category of Employee"],
            row["Grade / Service Details"], row["Year of Service"],
            row["Rate Type"], rate_val, row["Effective From"], row["Notes"],
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font, cell.fill, cell.alignment, cell.border = data_font, fill, data_align, border

# Rate column number format
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
    for cell in row:
        if isinstance(cell.value, float):
            cell.number_format = "#,##0.00"

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"

# ── Reference sheet ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Reference")
ref_hdr_fill = PatternFill("solid", fgColor="375623")
for col, h in enumerate(["Field", "Allowed Values / Format"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    c.fill  = ref_hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

ref_rows = [
    ("Sector",                    "Free text — must be spelled identically on every row for the same sector"),
    ("Category of Employee",      "Free text — e.g. Supervisor, General Worker, Cook"),
    ("Grade / Service Details",   "Optional — sub-grade description, leave blank if none"),
    ("Year of Service",           "Optional — e.g. 1st year / 2nd year / 5th year & above (leave blank if not year-based)"),
    ("Rate Type",                 "Exactly one of:  monthly  |  daily  |  hourly  |  per_kg  |  per_show"),
    ("Rate",                      "Number only — no commas, symbols or currency code — e.g. 17101 or 810.89"),
    ("Effective From",            "Date in YYYY-MM-DD format — e.g. 2024-01-01\n(Use 2024-07-01 for sectors with a July fiscal-year start, e.g. Sugar, Tea)"),
    ("Notes",                     "Optional free text — e.g. \"6-day week\", \"Per show above 28 shows\""),
]
bold10 = Font(name="Calibri", size=10, bold=True)
reg10  = Font(name="Calibri", size=10)
wrap   = Alignment(wrap_text=True, vertical="top")

for i, (field, desc) in enumerate(ref_rows, 2):
    c1 = ws2.cell(row=i, column=1, value=field)
    c1.font, c1.alignment = bold10, Alignment(vertical="top")
    c2 = ws2.cell(row=i, column=2, value=desc)
    c2.font, c2.alignment = reg10, wrap
    ws2.row_dimensions[i].height = 36

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 70
ws2.row_dimensions[1].height = 24

wb.save(out_path)
print(f"Saved: {out_path}")
